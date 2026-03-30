"""Vues de gestion des parcelles."""
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.views.generic import ListView, DetailView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db.models import Q
from django.http import JsonResponse
from django.core.serializers import serialize

from .models import Parcelle, ParcelleMedia, Zone
from .forms import ParcelleForm, ParcelleMediaForm, ParcelleSearchForm, BulkParcelleForm
from accounts.models import AccessLog
from accounts.decorators import role_required, vendeur_required, geometre_required


class ParcelleListView(ListView):
    """Liste des parcelles avec filtrage."""
    model = Parcelle
    template_name = "parcelles/parcelle_list.html"
    context_object_name = "parcelles"
    paginate_by = 12

    def get_queryset(self):
        user = self.request.user
        qs = Parcelle.objects.select_related("owner", "zone").prefetch_related("medias")

        # Admin/staff voient tout, vendeur voit ses propres parcelles non validées
        if user.is_authenticated and (user.is_admin_role or user.is_staff):
            pass  # Voir tout
        elif user.is_authenticated and user.is_vendeur:
            qs = qs.filter(Q(is_validated=True) | Q(owner=user))
        else:
            qs = qs.filter(is_validated=True)

        form = ParcelleSearchForm(self.request.GET)

        if form.is_valid():
            q = form.cleaned_data.get("q")
            if q:
                qs = qs.filter(
                    Q(title__icontains=q) | Q(lot_number__icontains=q) |
                    Q(address__icontains=q) | Q(description__icontains=q)
                )
            if form.cleaned_data.get("status"):
                qs = qs.filter(status=form.cleaned_data["status"])
            if form.cleaned_data.get("land_type"):
                qs = qs.filter(land_type=form.cleaned_data["land_type"])
            if form.cleaned_data.get("price_min"):
                qs = qs.filter(price__gte=form.cleaned_data["price_min"])
            if form.cleaned_data.get("price_max"):
                qs = qs.filter(price__lte=form.cleaned_data["price_max"])
            if form.cleaned_data.get("surface_min"):
                qs = qs.filter(surface_m2__gte=form.cleaned_data["surface_min"])
            if form.cleaned_data.get("surface_max"):
                qs = qs.filter(surface_m2__lte=form.cleaned_data["surface_max"])
            if form.cleaned_data.get("zone"):
                zone_obj = form.cleaned_data["zone"]
                zone_pk = zone_obj.pk if hasattr(zone_obj, 'pk') else zone_obj
                qs = qs.filter(
                    Q(zone_id=zone_pk) | Q(ilot__zone_id=zone_pk)
                )

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["form"] = ParcelleSearchForm(self.request.GET)
        ctx["zones"] = Zone.objects.all()
        ctx["total"] = self.get_queryset().count()
        # IDs des parcelles likées par l'utilisateur connecté
        if self.request.user.is_authenticated:
            from parcelles.models import ParcelleReaction
            ctx["user_liked_ids"] = set(
                ParcelleReaction.objects.filter(
                    user=self.request.user, reaction_type="like"
                ).values_list("parcelle_id", flat=True)
            )
        else:
            ctx["user_liked_ids"] = set()
        # IDs des parcelles avec promotion active
        from parcelles.models import PromotionCampaign
        ctx["promoted_ids"] = set(
            PromotionCampaign.objects.filter(
                status="active"
            ).values_list("parcelle_id", flat=True)
        )
        ctx["premium_ids"] = set(
            PromotionCampaign.objects.filter(
                status="active", campaign_type__in=["premium", "boost"]
            ).values_list("parcelle_id", flat=True)
        )
        return ctx


class ParcelleDetailView(DetailView):
    """Détail d'une parcelle avec médias, bons de visite, certification."""
    model = Parcelle
    template_name = "parcelles/parcelle_detail.html"
    context_object_name = "parcelle"

    def get_queryset(self):
        return Parcelle.objects.select_related("owner", "owner__profile", "zone", "ilot")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        parcelle = self.object
        user = self.request.user
        medias = parcelle.medias.all()
        ctx["medias"] = medias
        ctx["is_owner"] = user.is_authenticated and (parcelle.owner == user or user.is_staff)
        ctx["has_video"] = medias.filter(media_type__in=["video", "drone"]).exists()
        ctx["has_plan"] = medias.filter(media_type="plan").exists()

        # ── COTATION : vérifier si l'acheteur a payé sa cotation ──
        has_cotation = False
        cotation = None
        if user.is_authenticated and user.is_acheteur:
            from transactions.cotation_service import check_cotation_access
            cotation = check_cotation_access(user, parcelle)
            has_cotation = cotation is not None and cotation.is_valid

        ctx["has_cotation"] = has_cotation
        ctx["cotation"] = cotation
        ctx["cotation_amount"] = None
        if user.is_authenticated and user.is_acheteur and parcelle.price:
            from transactions.cotation_models import Cotation
            ctx["cotation_amount"] = Cotation.compute_cotation_amount(parcelle.price)

        # Documents visibles selon le rôle ET la cotation
        if user.is_authenticated:
            if user.is_staff or user == parcelle.owner:
                ctx["documents"] = parcelle.documents.all()
            elif user.is_acheteur and has_cotation:
                # Cotation payée → accès aux docs filigranés (buyer_only)
                ctx["documents"] = parcelle.documents.filter(
                    confidentiality__in=["public", "buyer_only"]
                )
            elif user.is_acheteur:
                # Pas de cotation → docs publics seulement
                ctx["documents"] = parcelle.documents.filter(confidentiality="public")
            else:
                ctx["documents"] = parcelle.documents.filter(confidentiality="public")
        else:
            ctx["documents"] = parcelle.documents.filter(confidentiality="public")

        # Bons de visite (acheteur connecté + cotation)
        if user.is_authenticated and user.is_acheteur:
            from transactions.models import BonDeVisite
            ctx["user_visits"] = BonDeVisite.objects.filter(
                visitor=user, parcelle=parcelle
            ).order_by("-created_at")[:3]
            # Visite possible seulement si cotation validée
            ctx["can_visit"] = parcelle.status == "disponible" and has_cotation
        else:
            ctx["user_visits"] = []
            ctx["can_visit"] = False

        # Badge certification du vendeur + avis vendeur
        if parcelle.owner:
            from accounts.models import CertificationRequest
            cert = CertificationRequest.objects.filter(
                user=parcelle.owner, status="approved"
            ).first()
            ctx["seller_certified"] = cert is not None

            # Notation du vendeur (via sa boutique)
            from transactions.cotation_models import Boutique, Review
            from django.db.models import Avg, Count
            try:
                boutique = Boutique.objects.get(owner=parcelle.owner, status="active")
                ctx["seller_boutique"] = boutique
                seller_reviews = Review.objects.filter(
                    boutique=boutique, is_visible=True,
                )
                stats = seller_reviews.aggregate(
                    avg_score=Avg("score"),
                    total=Count("id"),
                )
                ctx["seller_rating"] = stats["avg_score"]
                ctx["seller_review_count"] = stats["total"]
            except Boutique.DoesNotExist:
                ctx["seller_boutique"] = None
                ctx["seller_rating"] = None
                ctx["seller_review_count"] = 0
        else:
            ctx["seller_certified"] = False
            ctx["seller_boutique"] = None
            ctx["seller_rating"] = None
            ctx["seller_review_count"] = 0

        # Incrémenter les vues
        Parcelle.objects.filter(pk=parcelle.pk).update(views_count=parcelle.views_count + 1)

        # Log d'accès
        if user.is_authenticated:
            AccessLog.objects.create(
                user=user,
                action=AccessLog.ActionType.VIEW_PARCELLE,
                resource_type="Parcelle",
                resource_id=str(parcelle.pk),
                ip_address=_get_ip(self.request),
            )

        return ctx


@login_required
def parcelle_create_view(request):
    """Création d'une parcelle.

    Autorisé pour :
      - SuperUser (admin)
      - Vendeur / Promoteur avec une boutique active (cotation boutique payée)
    """
    user = request.user

    # Vérifier les droits
    has_boutique = False
    if user.is_vendeur or user.is_promoteur:
        boutique = getattr(user, "boutique", None)
        has_boutique = boutique is not None and boutique.is_active

    if not user.is_superuser and not has_boutique:
        if user.is_vendeur or user.is_promoteur:
            messages.warning(
                request,
                "Vous devez d'abord créer votre boutique (cotation boutique) "
                "avant de pouvoir publier des parcelles."
            )
            return redirect("transactions:boutique_cotation")
        messages.error(request, "Seuls les vendeurs et promoteurs avec une boutique active peuvent déposer des parcelles.")
        return redirect("parcelles:list")

    if request.method == "POST":
        form = ParcelleForm(request.POST, request.FILES)
        if form.is_valid():
            parcelle = form.save(commit=False)
            parcelle.owner = request.user
            # Auto-validation pour les superusers uniquement
            if request.user.is_superuser:
                parcelle.is_validated = True
                parcelle.validated_by = request.user
                from django.utils import timezone
                parcelle.validated_at = timezone.now()
            parcelle.save()

            # Incrémenter le compteur boutique
            if has_boutique:
                boutique = user.boutique
                boutique.total_parcelles = user.parcelles.count()
                boutique.save(update_fields=["total_parcelles"])

            # Invalider le cache GeoJSON immédiatement
            from parcelles.signals import invalidate_geojson_cache
            invalidate_geojson_cache()

            # Notifier les géomètres/admins qu'une parcelle attend validation
            if not parcelle.is_validated:
                try:
                    from notifications.services import send_notification
                    from accounts.models import User as UserModel
                    reviewers = UserModel.objects.filter(
                        Q(role="geometre") | Q(role="admin") | Q(is_staff=True)
                    ).distinct()
                    for reviewer in reviewers:
                        send_notification(
                            recipient=reviewer,
                            notification_type="system",
                            title="Nouvelle parcelle à valider",
                            message=f"La parcelle « {parcelle.title} » (Lot {parcelle.lot_number}) déposée par {user.get_full_name() or user.username} attend votre validation.",
                            data={"parcelle_id": str(parcelle.pk)},
                        )
                except Exception:
                    pass

            if parcelle.is_validated:
                messages.success(
                    request,
                    f"Parcelle \u00ab {parcelle.title} \u00bb creee et validee ! "
                    "Elle est visible sur la carte publique."
                )
            else:
                messages.success(
                    request,
                    f"Parcelle \u00ab {parcelle.title} \u00bb creee avec succes ! "
                    "Elle sera visible sur la carte apres validation par un geometre."
                )
            # Rediriger vers la carte focalisée sur la nouvelle parcelle
            return redirect(f"/carte/?focus={parcelle.pk}")
    else:
        form = ParcelleForm()

    return render(request, "parcelles/parcelle_form.html", {
        "form": form,
        "title": "Déposer une parcelle",
    })


@login_required
def parcelle_edit_view(request, pk):
    """Modification d'une parcelle par son propriétaire."""
    parcelle = get_object_or_404(Parcelle, pk=pk)
    if parcelle.owner != request.user and not request.user.is_superuser:
        messages.error(request, "Vous n'êtes pas autorisé à modifier cette parcelle.")
        return redirect("parcelles:detail", pk=pk)

    if request.method == "POST":
        form = ParcelleForm(request.POST, request.FILES, instance=parcelle)
        if form.is_valid():
            form.save()
            # Invalider le cache GeoJSON après modification
            from parcelles.signals import invalidate_geojson_cache
            invalidate_geojson_cache()
            messages.success(request, "Parcelle mise à jour.")
            return redirect("parcelles:detail", pk=pk)
    else:
        form = ParcelleForm(instance=parcelle)

    return render(request, "parcelles/parcelle_form.html", {
        "form": form,
        "parcelle": parcelle,
        "title": "Modifier la parcelle",
    })


@login_required
def media_upload_view(request, pk):
    """Upload de médias sur une parcelle."""
    parcelle = get_object_or_404(Parcelle, pk=pk)
    if parcelle.owner != request.user and not request.user.is_admin_role:
        messages.error(request, "Non autorisé.")
        return redirect("parcelles:detail", pk=pk)

    if request.method == "POST":
        form = ParcelleMediaForm(request.POST, request.FILES)
        if form.is_valid():
            media = form.save(commit=False)
            media.parcelle = parcelle
            media.save()

            # Appliquer le filigrane EYE-FONCIER sur les images
            if media.media_type in ("image", "plan"):
                try:
                    from parcelles.watermark_service import apply_watermark
                    import logging
                    wm_logger = logging.getLogger("parcelles.watermark_service")
                    file_path = media.file.path
                    wm_logger.info("Watermark upload : %s", file_path)
                    apply_watermark(file_path)
                except Exception as e:
                    import logging
                    logging.getLogger("parcelles.watermark_service").error(
                        "Erreur watermark upload : %s", e, exc_info=True
                    )

            messages.success(request, "Média ajouté avec succès.")
            return redirect("parcelles:media_upload", pk=pk)
    else:
        form = ParcelleMediaForm()

    return render(request, "parcelles/media_upload.html", {
        "form": form,
        "parcelle": parcelle,
        "medias": parcelle.medias.all(),
    })


@login_required
def media_delete_view(request, pk, media_pk):
    """Suppression d'un média d'une parcelle."""
    parcelle = get_object_or_404(Parcelle, pk=pk)
    media = get_object_or_404(ParcelleMedia, pk=media_pk, parcelle=parcelle)

    if parcelle.owner != request.user and not request.user.is_admin_role:
        messages.error(request, "Non autorisé à supprimer ce média.")
        return redirect("parcelles:detail", pk=pk)

    if request.method == "POST":
        title = media.title or media.get_media_type_display()
        # Delete the file from storage
        if media.file:
            media.file.delete(save=False)
        if media.thumbnail:
            media.thumbnail.delete(save=False)
        media.delete()
        messages.success(request, f"Média « {title} » supprimé avec succès.")

    # Redirect to the referer page if available, otherwise to detail
    referer = request.META.get("HTTP_REFERER", "")
    if "medias" in referer:
        return redirect("parcelles:media_upload", pk=pk)
    return redirect("parcelles:detail", pk=pk)


# ─── Validation Géomètre ───────────────────────────────
@geometre_required
def validate_parcelle_view(request, pk):
    """Validation technique d'une parcelle par un géomètre."""

    parcelle = get_object_or_404(Parcelle, pk=pk)

    if request.method == "POST":
        from django.utils import timezone
        parcelle.is_validated = True
        parcelle.validated_by = request.user
        parcelle.validated_at = timezone.now()

        # Garantir que le centroide est calcule (necessaire pour la carte)
        if parcelle.geometry and not parcelle.centroid:
            parcelle.centroid = parcelle.geometry.centroid

        parcelle.save()

        # Invalider le cache — la parcelle devient visible sur la carte publique
        from parcelles.signals import invalidate_geojson_cache
        invalidate_geojson_cache()

        # Forcer une double invalidation pour les caches distribues
        from django.core.cache import cache
        cache.delete_pattern("geojson:*") if hasattr(cache, 'delete_pattern') else None

        # Notifier le propriétaire que sa parcelle est validée
        try:
            from notifications.services import send_notification
            send_notification(
                recipient=parcelle.owner,
                notification_type="transaction_status",
                title="Parcelle validée !",
                message=f"Votre parcelle « {parcelle.title} » (Lot {parcelle.lot_number}) a été validée par un géomètre. Elle est désormais visible sur la carte publique.",
                data={"parcelle_id": str(parcelle.pk)},
            )
        except Exception:
            pass

        messages.success(request, f"Parcelle {parcelle.lot_number} validée avec succès.")
        return redirect("accounts:dashboard")

    return render(request, "parcelles/validate.html", {"parcelle": parcelle})


# ─── Suppression ───────────────────────────────────────
class ParcelleDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Parcelle
    template_name = "parcelles/parcelle_confirm_delete.html"
    success_url = reverse_lazy("accounts:dashboard")

    def test_func(self):
        parcelle = self.get_object()
        return parcelle.owner == self.request.user or self.request.user.is_admin_role


def _get_ip(request):
    x_forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
    return x_forwarded.split(",")[0].strip() if x_forwarded else request.META.get("REMOTE_ADDR")


# ═══════════════════════════════════════════════════════════
# RÉACTIONS SUR LES PARCELLES
# ═══════════════════════════════════════════════════════════

@login_required
def toggle_reaction_view(request, pk):
    """Toggle une réaction sur une parcelle (AJAX ou POST)."""
    from .models import ParcelleReaction

    parcelle = get_object_or_404(Parcelle, pk=pk)
    reaction_type = request.POST.get("reaction_type", "like")

    if reaction_type not in dict(ParcelleReaction.ReactionType.choices):
        return JsonResponse({"error": "Type de réaction invalide."}, status=400)

    reaction, created = ParcelleReaction.objects.get_or_create(
        user=request.user,
        parcelle=parcelle,
        reaction_type=reaction_type,
    )

    if not created:
        reaction.delete()
        active = False
    else:
        active = True

    # Compter les réactions
    counts = {}
    for rt in ParcelleReaction.ReactionType.values:
        counts[rt] = parcelle.reactions.filter(reaction_type=rt).count()

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({
            "active": active,
            "reaction_type": reaction_type,
            "counts": counts,
        })

    messages.success(
        request,
        "Réaction {} !".format("ajoutée" if active else "retirée"),
    )
    return redirect("parcelles:detail", pk=pk)


def parcelle_reactions_api(request, pk):
    """API publique : réactions d'une parcelle."""
    from .models import ParcelleReaction

    parcelle = get_object_or_404(Parcelle, pk=pk)

    counts = {}
    for rt in ParcelleReaction.ReactionType.values:
        counts[rt] = parcelle.reactions.filter(reaction_type=rt).count()

    user_reactions = []
    if request.user.is_authenticated:
        user_reactions = list(
            parcelle.reactions.filter(user=request.user).values_list("reaction_type", flat=True)
        )

    return JsonResponse({
        "parcelle_id": str(pk),
        "counts": counts,
        "user_reactions": user_reactions,
    })


# ═══════════════════════════════════════════════════════════
# PARCELLES À PROXIMITÉ (géolocalisation)
# ═══════════════════════════════════════════════════════════

def nearby_parcelles_api(request):
    """API : parcelles à proximité d'une position GPS.

    Paramètres GET :
      - lat : latitude
      - lng : longitude
      - radius : rayon en mètres (défaut 5000, max 50000)
      - limit : nombre max de résultats (défaut 20)
    """
    from django.contrib.gis.geos import Point
    from django.contrib.gis.measure import D
    from django.contrib.gis.db.models.functions import Distance

    try:
        lat = float(request.GET.get("lat", 0))
        lng = float(request.GET.get("lng", 0))
        radius = min(int(request.GET.get("radius", 5000)), 50000)
        limit = min(int(request.GET.get("limit", 20)), 50)
    except (ValueError, TypeError):
        return JsonResponse({"error": "Paramètres invalides."}, status=400)

    if lat == 0 and lng == 0:
        return JsonResponse({"error": "Coordonnées manquantes (lat, lng)."}, status=400)

    user_point = Point(lng, lat, srid=4326)

    parcelles = Parcelle.objects.filter(
        status="disponible",
        is_validated=True,
        centroid__isnull=False,
        centroid__distance_lte=(user_point, D(m=radius)),
    ).annotate(
        distance=Distance("centroid", user_point),
    ).select_related("zone", "owner").order_by("distance")[:limit]

    results = []
    for p in parcelles:
        results.append({
            "id": str(p.pk),
            "lot_number": p.lot_number,
            "title": p.title,
            "price": float(p.price) if p.price else 0,
            "surface_m2": float(p.surface_m2) if p.surface_m2 else 0,
            "land_type": p.get_land_type_display(),
            "zone": p.zone.name if p.zone else "",
            "distance_m": round(p.distance.m, 0) if p.distance else None,
            "lat": p.centroid.y,
            "lng": p.centroid.x,
            "status": p.status,
            "owner": p.owner.get_full_name() if p.owner else "",
            "trust_badge": p.trust_badge,
            "is_validated": p.is_validated,
            "main_image": p.main_image,
        })

    return JsonResponse({
        "center": {"lat": lat, "lng": lng},
        "radius_m": radius,
        "count": len(results),
        "parcelles": results,
    })


# ═══════════════════════════════════════════════════════════
# PROMOTIONS VENDEUR
# ═══════════════════════════════════════════════════════════

@login_required
def promotion_create_view(request, pk):
    """Création d'une campagne de promotion pour une parcelle."""
    from .models import PromotionCampaign

    parcelle = get_object_or_404(Parcelle, pk=pk)

    # Seul le propriétaire peut promouvoir
    if request.user != parcelle.owner and not request.user.is_staff:
        messages.error(request, "Seul le propriétaire peut promouvoir cette parcelle.")
        return redirect("parcelles:detail", pk=pk)

    if request.method == "POST":
        campaign_type = request.POST.get("campaign_type", "basic")
        duration_weeks = int(request.POST.get("duration_weeks", 1))
        highlight_text = request.POST.get("highlight_text", "")
        payment_method = request.POST.get("payment_method", "mobile_money")

        if campaign_type not in dict(PromotionCampaign.CampaignType.choices):
            messages.error(request, "Type de campagne invalide.")
            return redirect("parcelles:promote", pk=pk)

        duration_weeks = max(1, min(duration_weeks, 52))

        campaign = PromotionCampaign.objects.create(
            parcelle=parcelle,
            seller=request.user,
            campaign_type=campaign_type,
            duration_weeks=duration_weeks,
            highlight_text=highlight_text,
            payment_method=payment_method,
            amount_paid=PromotionCampaign.PRICING.get(campaign_type, 5000) * duration_weeks,
            status="pending_payment",
        )

        # Rediriger vers le paiement en ligne
        from django.http import QueryDict
        payment_data = {
            "payment_type": "promotion",
            "amount": str(campaign.total_price),
            "reference_id": str(campaign.pk),
            "description": "Promotion {} — Lot {} ({} sem.)".format(
                campaign.get_campaign_type_display(),
                parcelle.lot_number,
                duration_weeks,
            ),
        }
        request.session["pending_promotion_payment"] = payment_data

        # Construction du POST vers payment_initiate
        from django.test import RequestFactory
        messages.info(
            request,
            "Campagne {} créée ! Procédez au paiement de {:,.0f} FCFA.".format(
                campaign.get_campaign_type_display(),
                float(campaign.amount_paid),
            ),
        )
        # Stocker pour redirection
        request.session["auto_payment"] = payment_data
        return redirect("parcelles:promotion_detail", pk=pk, promo_pk=campaign.pk)

    # GET — page de création
    context = {
        "parcelle": parcelle,
        "pricing": PromotionCampaign.PRICING,
        "campaign_types": PromotionCampaign.CampaignType.choices,
    }
    return render(request, "parcelles/promotion_create.html", context)


@login_required
def promotion_detail_view(request, pk, promo_pk):
    """Détail d'une campagne de promotion."""
    from .models import PromotionCampaign

    parcelle = get_object_or_404(Parcelle, pk=pk)
    campaign = get_object_or_404(PromotionCampaign, pk=promo_pk, parcelle=parcelle)

    context = {"parcelle": parcelle, "campaign": campaign}
    return render(request, "parcelles/promotion_detail.html", context)


@login_required
def my_promotions_view(request):
    """Liste des promotions du vendeur connecté."""
    from .models import PromotionCampaign

    campaigns = PromotionCampaign.objects.filter(
        seller=request.user,
    ).select_related("parcelle", "parcelle__zone").order_by("-created_at")

    context = {"campaigns": campaigns}
    return render(request, "parcelles/my_promotions.html", context)


def promoted_parcelles_api(request):
    """API : parcelles promues, adaptées au profil de l'acheteur.
    Utilise le Smart Matching si un profil acheteur existe.
    """
    from .models import PromotionCampaign
    from django.utils import timezone

    now = timezone.now()

    # Campagnes actives
    active_campaigns = PromotionCampaign.objects.filter(
        status="active",
        start_date__lte=now,
        end_date__gte=now,
        parcelle__status="disponible",
    ).select_related("parcelle", "parcelle__zone", "parcelle__owner")

    results = []
    user = request.user

    for campaign in active_campaigns:
        p = campaign.parcelle
        score = 50  # score par défaut

        # Smart Matching si profil acheteur
        if user.is_authenticated:
            try:
                from analysis.models import MatchScore
                ms = MatchScore.objects.filter(
                    buyer_profile__user=user, parcelle=p,
                ).first()
                if ms:
                    score = ms.final_score
            except Exception:
                pass

        # Incrémenter les impressions
        campaign.impressions += 1
        campaign.save(update_fields=["impressions"])

        results.append({
            "id": str(p.pk),
            "lot_number": p.lot_number,
            "title": p.title,
            "price": float(p.price) if p.price else 0,
            "surface_m2": float(p.surface_m2) if p.surface_m2 else 0,
            "zone": p.zone.name if p.zone else "",
            "campaign_type": campaign.campaign_type,
            "highlight": campaign.highlight_text,
            "match_score": round(score, 1),
            "lat": p.centroid.y if p.centroid else None,
            "lng": p.centroid.x if p.centroid else None,
        })

    # Trier par score de matching (les plus pertinents en premier)
    results.sort(key=lambda x: x["match_score"], reverse=True)

    return JsonResponse({"promoted": results, "count": len(results)})


def recommended_parcelles_api(request):
    """API : recommandations personnalisées pour le popup promotion.
    Utilise le service de recommandation multi-stratégie.
    """
    from .recommendation_service import get_personalized_recommendations

    user = request.user if request.user.is_authenticated else request.user
    limit = int(request.GET.get("limit", 5))
    recommendations = get_personalized_recommendations(user, limit=limit)
    return JsonResponse({"recommendations": recommendations, "count": len(recommendations)})


# ═══════════════════════════════════════════════════════
# Phase 9 : Mode Pro Promoteur
# ═══════════════════════════════════════════════════════

@vendeur_required
def bulk_parcelle_upload_view(request):
    """Import en lot de parcelles depuis un fichier CSV/Excel."""

    results = None
    if request.method == "POST":
        form = BulkParcelleForm(request.POST, request.FILES)
        if form.is_valid():
            results = _process_bulk_file(request.user, form.cleaned_data["file"])
            if results["created"]:
                messages.success(request, f"{results['created']} parcelle(s) créée(s) avec succès.")
            if results["errors"]:
                messages.warning(request, f"{len(results['errors'])} ligne(s) en erreur.")
    else:
        form = BulkParcelleForm()

    return render(request, "parcelles/bulk_upload.html", {
        "form": form,
        "results": results,
    })


def _process_bulk_file(owner, uploaded_file):
    """Parse et crée les parcelles depuis un CSV/Excel."""
    import csv
    import io
    import os
    from decimal import Decimal, InvalidOperation

    results = {"created": 0, "errors": [], "total": 0}
    ext = os.path.splitext(uploaded_file.name)[1].lower()

    rows = []
    if ext == ".csv":
        content = uploaded_file.read().decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(content), delimiter=";")
        # Essayer avec virgule si pas de colonnes
        fieldnames = reader.fieldnames or []
        if len(fieldnames) <= 1:
            reader = csv.DictReader(io.StringIO(content), delimiter=",")
        for row in reader:
            rows.append(row)
    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(uploaded_file, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            headers = [h.strip().lower() if h else "" for h in headers]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        except ImportError:
            results["errors"].append({"line": 0, "error": "Module openpyxl non installé pour lire les fichiers Excel."})
            return results

    REQUIRED = ["lot_number"]
    VALID_LAND_TYPES = dict(Parcelle.LandType.choices).keys() if hasattr(Parcelle, "LandType") else []

    for i, row in enumerate(rows, start=2):
        results["total"] += 1
        # Normaliser les clés
        row = {k.strip().lower(): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}

        lot = row.get("lot_number", "")
        if not lot:
            results["errors"].append({"line": i, "error": "lot_number manquant"})
            continue

        # Vérifier doublon
        if Parcelle.objects.filter(lot_number=lot, owner=owner).exists():
            results["errors"].append({"line": i, "error": f"Lot {lot} existe déjà"})
            continue

        try:
            surface = Decimal(str(row.get("surface_m2", 0) or 0))
            price = Decimal(str(row.get("price", 0) or 0))
        except (InvalidOperation, ValueError):
            results["errors"].append({"line": i, "error": f"Lot {lot} — surface ou prix invalide"})
            continue

        land_type = row.get("land_type", "residential")
        if VALID_LAND_TYPES and land_type not in VALID_LAND_TYPES:
            land_type = "residential"

        try:
            Parcelle.objects.create(
                owner=owner,
                lot_number=lot,
                title=row.get("title", f"Lot {lot}") or f"Lot {lot}",
                description=row.get("description", "") or "",
                land_type=land_type,
                surface_m2=surface,
                price=price,
                address=row.get("address", "") or "",
                status="disponible",
            )
            results["created"] += 1
        except Exception as e:
            results["errors"].append({"line": i, "error": f"Lot {lot} — {str(e)[:80]}"})

    return results


@vendeur_required
def seller_dashboard_pro_view(request):
    """Dashboard promoteur avec stats avancées et gestion en lot."""

    parcelles = request.user.parcelles.all()
    from transactions.models import Transaction
    transactions = Transaction.objects.filter(seller=request.user)

    # Actions groupées
    if request.method == "POST":
        action = request.POST.get("bulk_action")
        selected_ids = request.POST.getlist("selected_parcelles")
        if selected_ids and action:
            selected = parcelles.filter(pk__in=selected_ids)
            if action == "set_price":
                new_price = request.POST.get("new_price")
                if new_price:
                    from decimal import Decimal
                    selected.update(price=Decimal(new_price))
                    messages.success(request, f"Prix mis à jour pour {selected.count()} parcelle(s).")
            elif action == "set_disponible":
                selected.update(status="disponible")
                messages.success(request, f"{selected.count()} parcelle(s) remise(s) en disponible.")

    # Stats
    from django.db.models import Sum, Count
    stats = {
        "total": parcelles.count(),
        "disponible": parcelles.filter(status="disponible").count(),
        "reserve": parcelles.filter(status="reserve").count(),
        "vendu": parcelles.filter(status="vendu").count(),
        "total_revenue": transactions.filter(status="completed").aggregate(s=Sum("amount"))["s"] or 0,
        "transactions_pending": transactions.filter(status__in=["pending", "reserved", "escrow_funded"]).count(),
    }

    return render(request, "parcelles/seller_dashboard_pro.html", {
        "parcelles": parcelles,
        "stats": stats,
    })


# ═══════════════════════════════════════════════════════════
# ANALYSE FONCIÈRE
# ═══════════════════════════════════════════════════════════

@login_required
def parcelle_analysis_view(request, pk):
    """Affiche le rapport d'analyse foncière d'une parcelle."""
    parcelle = get_object_or_404(
        Parcelle.objects.select_related("owner", "zone"),
        pk=pk,
    )
    user = request.user
    is_owner = user == parcelle.owner or user.is_staff
    is_verificateur = getattr(user, "is_geometre", False) or getattr(user, "is_admin_role", False) or user.is_staff

    # L'analyse est visible par tout utilisateur connecté (indicateur de confiance)
    # Seuls le propriétaire, vérificateurs et admins peuvent modifier/valider

    from .models import ParcelleAnalysis
    analysis = getattr(parcelle, "analysis", None)

    # Score breakdown pour le radar chart foncier (0-100)
    score_breakdown = None
    if analysis:
        score_breakdown = json.dumps({
            "labels": ["Géométrie", "Documents", "Chevauchement", "Terrain", "Propriété"],
            "scores": [
                analysis.score_geometry,
                analysis.score_documents,
                analysis.score_overlap,
                analysis.score_terrain,
                analysis.score_ownership,
            ],
        })

    # ═══ FUSION : données du module analysis (topographie, risques, proximité) ═══
    terrain = None
    risk = None
    constraints = []
    proximities = []
    radar_data_sig = None
    reports = []

    try:
        from analysis.models import (
            TerrainAnalysis, RiskAssessment,
            SpatialConstraint, ProximityAnalysis, AnalysisReport,
        )
        terrain = getattr(parcelle, "terrain_analysis", None)
        risk = getattr(parcelle, "risk_assessment", None)
        constraints = list(SpatialConstraint.objects.filter(parcelle=parcelle))
        proximities = list(ProximityAnalysis.objects.filter(parcelle=parcelle))
        reports = list(
            AnalysisReport.objects.filter(parcelle=parcelle, status="ready")
            .order_by("-generated_at")[:5]
        )

        # Radar SIG (5 axes — 0 à 5)
        if risk:
            radar_data_sig = json.dumps({
                "Accessibilité": float(risk.score_accessibility),
                "Topographie": float(risk.score_topography),
                "Juridique": float(risk.score_legal),
                "Environnement": float(risk.score_environment),
                "Prix": float(risk.score_price),
            })
    except ImportError:
        pass  # Module analysis non disponible

    is_analyzed_sig = terrain is not None

    # ── Validation par un vérificateur ──
    if request.method == "POST" and is_verificateur and analysis:
        action = request.POST.get("action")
        notes = request.POST.get("notes", "")
        if action == "validate":
            from .analysis_service import validate_analysis
            validate_analysis(analysis, user, notes)
            messages.success(request, f"Analyse validée — score {analysis.overall_score}/100 ({analysis.reliability_label}).")
        elif action == "reject":
            from .analysis_service import validate_analysis
            analysis.overall_score = min(analysis.overall_score, 39)  # Force rejet
            validate_analysis(analysis, user, notes)
            messages.warning(request, "Analyse rejetée.")
        return redirect("parcelles:analysis", pk=pk)

    return render(request, "parcelles/parcelle_analysis.html", {
        "parcelle": parcelle,
        "analysis": analysis,
        "is_owner": is_owner,
        "is_verificateur": is_verificateur,
        "score_breakdown": score_breakdown,
        # Données module analysis (SIG)
        "terrain": terrain,
        "risk": risk,
        "constraints": constraints,
        "proximities": proximities,
        "radar_data_sig": radar_data_sig,
        "is_analyzed_sig": is_analyzed_sig,
        "reports": reports,
    })


@login_required
def run_analysis_view(request, pk):
    """Lance l'analyse foncière automatisée d'une parcelle."""
    parcelle = get_object_or_404(Parcelle, pk=pk)
    user = request.user

    # Seul le propriétaire, un vérificateur ou un admin peut lancer l'analyse
    is_allowed = (
        user == parcelle.owner
        or getattr(user, "is_geometre", False)
        or getattr(user, "is_admin_role", False)
        or user.is_staff
    )
    if not is_allowed:
        messages.warning(request, "Vous n'êtes pas autorisé à lancer cette analyse.")
        return redirect("parcelles:detail", pk=pk)

    if request.method == "POST":
        from .analysis_service import run_full_analysis

        # 1. Lancer d'abord l'analyse SIG (topographie, risques, proximité)
        sig_ok = False
        try:
            from analysis.services.terrain_analyzer import analyze_parcelle_complete
            sig_result = analyze_parcelle_complete(parcelle)
            sig_score = sig_result["risk_assessment"].overall_score or 0
            sig_ok = True
        except Exception as e:
            logger.warning("Analyse SIG non disponible: %s", e)

        # 2. Puis l'analyse foncière (géométrie, documents, chevauchement, propriété)
        try:
            analysis = run_full_analysis(parcelle, inspector=user)
            msg = f"Analyse foncière terminée — Score : {analysis.overall_score}/100 (Grade {analysis.reliability_grade})"
            if sig_ok:
                msg += f" | SIG : {sig_score:.0f}/100"
            messages.success(request, msg)
        except Exception as e:
            messages.error(request, f"Erreur lors de l'analyse : {e}")
        return redirect("parcelles:analysis", pk=pk)

    return redirect("parcelles:analysis", pk=pk)


@login_required
def terrain_inspection_view(request, pk):
    """Enregistre une inspection terrain par un vérificateur."""
    parcelle = get_object_or_404(Parcelle, pk=pk)
    user = request.user

    # Seuls les vérificateurs/géomètres/admin
    is_verificateur = getattr(user, "is_geometre", False) or getattr(user, "is_admin_role", False) or user.is_staff
    if not is_verificateur:
        messages.warning(request, "Seuls les vérificateurs peuvent effectuer une inspection terrain.")
        return redirect("parcelles:detail", pk=pk)

    from .models import ParcelleAnalysis
    analysis = getattr(parcelle, "analysis", None)
    if not analysis:
        messages.warning(request, "Veuillez d'abord lancer l'analyse automatisée.")
        return redirect("parcelles:analysis", pk=pk)

    if request.method == "POST":
        score = request.POST.get("terrain_score", "0")
        notes = request.POST.get("terrain_notes", "")
        try:
            score = max(0, min(100, int(score)))
        except (ValueError, TypeError):
            score = 0

        from .analysis_service import record_terrain_inspection
        record_terrain_inspection(analysis, inspector=user, score=score, notes=notes)
        messages.success(request, f"Inspection terrain enregistrée — Score terrain : {score}/100.")
        return redirect("parcelles:analysis", pk=pk)

    return render(request, "parcelles/terrain_inspection.html", {
        "parcelle": parcelle,
        "analysis": analysis,
    })


# ═══════════════════════════════════════════════════════════
# SOUMISSION POUR VALIDATION & TABLEAU DE BORD GÉOMÈTRE
# ═══════════════════════════════════════════════════════════

@login_required
def submit_for_validation_view(request, pk):
    """Un vendeur soumet sa parcelle pour validation par un géomètre."""
    parcelle = get_object_or_404(Parcelle, pk=pk)
    user = request.user

    if user != parcelle.owner and not user.is_staff:
        messages.warning(request, "Seul le propriétaire peut soumettre cette parcelle.")
        return redirect("parcelles:detail", pk=pk)

    if parcelle.is_validated:
        messages.info(request, "Cette parcelle est déjà validée.")
        return redirect("parcelles:detail", pk=pk)

    if request.method == "POST":
        from .analysis_service import submit_for_validation
        try:
            analysis = submit_for_validation(parcelle, requester=user)
            messages.success(
                request,
                f"Parcelle soumise pour validation. Score automatique : {analysis.overall_score}/100. "
                f"Les géomètres ont été notifiés."
            )
        except Exception as e:
            messages.error(request, f"Erreur lors de la soumission : {e}")
        return redirect("parcelles:analysis", pk=pk)

    return render(request, "parcelles/submit_validation.html", {"parcelle": parcelle})


@login_required
def geometre_dashboard_view(request):
    """Tableau de bord du géomètre — parcelles en attente de validation."""
    user = request.user
    if not (getattr(user, "is_geometre", False) or getattr(user, "is_admin_role", False) or user.is_staff):
        messages.warning(request, "Accès réservé aux géomètres.")
        return redirect("accounts:dashboard")

    from .models import ParcelleAnalysis
    from django.db.models import Count

    # Parcelles en attente de validation (analyse terminée mais pas validée)
    pending = ParcelleAnalysis.objects.filter(
        status=ParcelleAnalysis.AnalysisStatus.IN_PROGRESS,
    ).select_related("parcelle", "parcelle__owner", "parcelle__zone").order_by("-updated_at")

    # Parcelles en attente d'inspection terrain
    needs_inspection = ParcelleAnalysis.objects.filter(
        status=ParcelleAnalysis.AnalysisStatus.IN_PROGRESS,
        terrain_inspected=False,
    ).select_related("parcelle", "parcelle__owner").order_by("-updated_at")

    # Validations récentes du géomètre
    my_validations = ParcelleAnalysis.objects.filter(
        analyzed_by=user,
    ).select_related("parcelle").order_by("-validated_at")[:20]

    # Stats
    stats = {
        "pending_count": pending.count(),
        "needs_inspection_count": needs_inspection.count(),
        "my_total": ParcelleAnalysis.objects.filter(analyzed_by=user).count(),
        "my_validated": ParcelleAnalysis.objects.filter(
            analyzed_by=user, status=ParcelleAnalysis.AnalysisStatus.VALIDATED,
        ).count(),
        "my_rejected": ParcelleAnalysis.objects.filter(
            analyzed_by=user, status=ParcelleAnalysis.AnalysisStatus.REJECTED,
        ).count(),
    }

    return render(request, "parcelles/geometre_dashboard.html", {
        "pending": pending,
        "needs_inspection": needs_inspection,
        "my_validations": my_validations,
        "stats": stats,
    })


# ═══════════════════════════════════════════════════════════
# LOTISSEMENTS (Promoteurs)
# ═══════════════════════════════════════════════════════════

@login_required
def lotissement_list_view(request):
    """Liste des lotissements du promoteur connecté (ou tous pour admin)."""
    from .models import Lotissement
    user = request.user

    if user.is_staff or getattr(user, "is_admin_role", False):
        qs = Lotissement.objects.select_related("promoteur", "zone").all()
    elif getattr(user, "is_promoteur", False) or getattr(user, "is_vendeur", False):
        qs = Lotissement.objects.filter(promoteur=user).select_related("zone")
    else:
        qs = Lotissement.objects.filter(status="active").select_related("promoteur", "zone")

    return render(request, "parcelles/lotissement_list.html", {
        "lotissements": qs,
    })


@login_required
def lotissement_create_view(request):
    """Création d'un nouveau lotissement par un promoteur."""
    user = request.user
    if not (getattr(user, "is_promoteur", False) or getattr(user, "is_vendeur", False) or user.is_staff):
        messages.warning(request, "Seuls les promoteurs peuvent créer un lotissement.")
        return redirect("parcelles:lotissement_list")

    from .models import Lotissement

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        code = request.POST.get("code", "").strip()
        description = request.POST.get("description", "")
        geometry_json = request.POST.get("geometry_json", "")
        zone_id = request.POST.get("zone")
        total_ilots = request.POST.get("total_ilots", 0)
        price_min = request.POST.get("price_per_m2_min")
        price_max = request.POST.get("price_per_m2_max")

        errors = []
        if not name:
            errors.append("Le nom est requis.")
        if not code:
            errors.append("Le code est requis.")
        elif Lotissement.objects.filter(code=code).exists():
            errors.append(f"Le code « {code} » est déjà utilisé.")
        if not geometry_json:
            errors.append("Veuillez dessiner le périmètre du lotissement sur la carte.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            try:
                from django.contrib.gis.geos import GEOSGeometry
                geom = GEOSGeometry(geometry_json, srid=4326)
                if geom.geom_type == "MultiPolygon":
                    geom = geom[0]

                lot = Lotissement(
                    promoteur=user,
                    name=name,
                    code=code,
                    description=description,
                    geometry=geom,
                    total_ilots=int(total_ilots or 0),
                    has_water="has_water" in request.POST,
                    has_electricity="has_electricity" in request.POST,
                    has_road="has_road" in request.POST,
                    has_drainage="has_drainage" in request.POST,
                )
                if zone_id:
                    lot.zone_id = zone_id
                if price_min:
                    lot.price_per_m2_min = int(price_min)
                if price_max:
                    lot.price_per_m2_max = int(price_max)

                # Calculer la surface via projection UTM
                try:
                    geom_utm = geom.transform(32630, clone=True)
                    lot.total_surface_m2 = round(geom_utm.area, 2)
                except Exception:
                    lot.total_surface_m2 = 0

                lot.has_public_spaces = "has_public_spaces" in request.POST
                lot.approval_reference = request.POST.get("approval_reference", "").strip()
                lot.is_approved = bool(lot.approval_reference)

                lot.save()

                # Upload fichiers
                file_fields = ["plan_image", "arrete_approbation", "attestation_villageoise",
                               "dossier_technique", "certificat_propriete"]
                update_fields = []
                for field_name in file_fields:
                    uploaded = request.FILES.get(field_name)
                    if uploaded:
                        setattr(lot, field_name, uploaded)
                        update_fields.append(field_name)
                if update_fields:
                    lot.save(update_fields=update_fields)

                messages.success(request, f"Lotissement « {name} » créé avec succès.")
                return redirect("parcelles:lotissement_detail", pk=lot.pk)
            except Exception as e:
                messages.error(request, f"Erreur : {e}")

    zones = Zone.objects.all()
    return render(request, "parcelles/lotissement_form.html", {"zones": zones})


@login_required
def lotissement_detail_view(request, pk):
    """Détail d'un lotissement avec ses parcelles."""
    from .models import Lotissement
    lot = get_object_or_404(
        Lotissement.objects.select_related("promoteur", "zone"),
        pk=pk,
    )
    user = request.user
    is_owner = user == lot.promoteur or user.is_staff

    parcelles = lot.parcelles.select_related("owner").order_by("lot_number")

    # Stats du lotissement
    stats = {
        "total": parcelles.count(),
        "disponible": parcelles.filter(status="disponible").count(),
        "reserve": parcelles.filter(status="reserve").count(),
        "vendu": parcelles.filter(status="vendu").count(),
    }

    return render(request, "parcelles/lotissement_detail.html", {
        "lot": lot,
        "parcelles": parcelles,
        "stats": stats,
        "is_owner": is_owner,
    })


@login_required
def lotissement_add_parcelle_view(request, pk):
    """Ajoute une parcelle à un lotissement existant."""
    from .models import Lotissement
    from .forms import ParcelleForm
    lot = get_object_or_404(Lotissement, pk=pk)
    user = request.user

    if user != lot.promoteur and not user.is_staff:
        messages.warning(request, "Vous n'êtes pas le propriétaire de ce lotissement.")
        return redirect("parcelles:lotissement_detail", pk=pk)

    if request.method == "POST":
        form = ParcelleForm(request.POST, request.FILES)
        if form.is_valid():
            parcelle = form.save(commit=False)
            parcelle.owner = user
            parcelle.lotissement = lot
            if lot.zone:
                parcelle.zone = lot.zone
            parcelle.save()
            lot.update_counters()
            messages.success(request, f"Parcelle « {parcelle.lot_number} » ajoutée au lotissement.")
            return redirect("parcelles:lotissement_detail", pk=pk)
    else:
        initial = {"zone": lot.zone_id} if lot.zone else {}
        form = ParcelleForm(initial=initial)

    return render(request, "parcelles/parcelle_form.html", {
        "form": form,
        "parcelle": None,
        "lotissement": lot,
    })
